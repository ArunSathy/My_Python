import openpyxl

wb=openpyxl.load_workbook('C:\\Users\\arhsathy\\Desktop\\My Python file\\data.xlsx')

# print(type(wb))
# print(wb.sheetnames)

# print(wb.active.title)

x=wb['DATA']

# x1=x['A2'].value

# ways to print the datas from an excel ------------

# way 1 ------------

# x2=wb['DATA']['A2'].value
#
# print(x)

# way 2 ------------

print(x.cell(2,2).value)